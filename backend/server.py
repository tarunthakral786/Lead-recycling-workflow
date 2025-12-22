from fastapi import FastAPI, APIRouter, HTTPException, Depends, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication")

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get('name') != 'TT':
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# Models
class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: EmailStr
    hashed_password: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserResponse(BaseModel):
    id: str
    name: str
    email: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

class RecoverySettings(BaseModel):
    pp_battery_percent: float = 60.5
    mc_smf_battery_percent: float = 57.5
    hr_battery_percent: float = 50.0

class RefiningBatch(BaseModel):
    input_source: str = "manual"  # 'manual', 'SANTOSH', or RML SKU
    sb_percentage: Optional[float] = None  # Required when input_source is SANTOSH
    lead_ingot_kg: float
    lead_ingot_pieces: int
    lead_ingot_image: str
    initial_dross_kg: float
    initial_dross_image: str
    dross_2nd_kg: float
    dross_2nd_image: str
    dross_3rd_kg: float
    dross_3rd_image: str
    dross_remarks: Optional[str] = ""
    pure_lead_kg: float
    pure_lead_pieces: int = 0
    pure_lead_image: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RefiningEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    entry_type: str = "refining"
    batches: List[RefiningBatch]
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RecyclingBatch(BaseModel):
    battery_type: str
    battery_kg: float
    battery_image: str
    quantity_received: float
    remelted_lead_kg: float
    receivable_kg: float
    recovery_percent: float
    remelted_lead_image: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RecyclingEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    entry_type: str = "recycling"
    batches: List[RecyclingBatch]
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DrossRecyclingBatch(BaseModel):
    dross_type: str
    quantity_sent: float
    high_lead_recovered: float
    spectro_image: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DrossRecyclingEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    batches: List[DrossRecyclingBatch]
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SaleEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    party_name: str
    sku_type: str  # "Pure Lead", "High Lead", or RML SKU name
    quantity_kg: float
    entry_date: Optional[str] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SaleCreate(BaseModel):
    party_name: str
    sku_type: str  # "Pure Lead", "High Lead", or RML SKU name
    quantity_kg: float
    entry_date: Optional[str] = None

class SummaryStats(BaseModel):
    # New simplified dashboard stats
    pure_lead_stock: float = 0  # Pure Lead produced - Pure Lead sold
    rml_stock: float = 0  # RML purchased - RML used in refining - RML sold
    total_receivable: float = 0  # Recycling receivable - SANTOSH usage
    high_lead_stock: float = 0  # High Lead recovered - High Lead sold
    total_dross: float = 0  # Sum of all dross from refining
    antimony_recoverable: float = 0  # SB% x Quantity for each refining batch
    # Legacy fields for backward compatibility
    total_pure_lead_manufactured: float = 0
    total_remelted_lead: float = 0
    total_sold: float = 0
    available_stock: float = 0
    remelted_lead_in_stock: float = 0
    total_high_lead: float = 0
    total_rml_purchased: float = 0

class AvailableSKU(BaseModel):
    sku_type: str  # "Pure Lead", "High Lead", or RML SKU name
    sb_percentage: Optional[float] = None
    available_kg: float
    display_name: str

# RML Purchase Models
class RMLPurchaseBatch(BaseModel):
    quantity_kg: float
    pieces: int
    sb_percentage: float
    remarks: str = ""
    image: str
    sku: str = ""
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RMLPurchaseEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    entry_type: str = "rml_purchase"
    batches: List[RMLPurchaseBatch]
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Routes
@api_router.get("/")
async def root():
    return {"message": "SPES PRO API"}

# Admin - User Management
@api_router.post("/admin/users", response_model=UserResponse)
async def create_user(user_data: UserCreate, admin: dict = Depends(require_admin)):
    existing_user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        name=user_data.name,
        email=user_data.email,
        hashed_password=hash_password(user_data.password)
    )
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    return UserResponse(id=user.id, name=user.name, email=user.email)

@api_router.get("/admin/users", response_model=List[UserResponse])
async def get_all_users(admin: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "hashed_password": 0}).to_list(100)
    return [UserResponse(id=u['id'], name=u['name'], email=u['email']) for u in users]

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    # Prevent deleting TT account
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user and user.get('name') == 'TT':
        raise HTTPException(status_code=400, detail="Cannot delete TT admin account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

# Admin - Recovery Settings
@api_router.get("/admin/recovery-settings", response_model=RecoverySettings)
async def get_recovery_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"type": "recovery_settings"}, {"_id": 0})
    if not settings:
        return RecoverySettings()
    return RecoverySettings(**settings)

@api_router.put("/admin/recovery-settings")
async def update_recovery_settings(settings: RecoverySettings, admin: dict = Depends(require_admin)):
    await db.settings.update_one(
        {"type": "recovery_settings"},
        {"$set": {
            "pp_battery_percent": settings.pp_battery_percent,
            "mc_smf_battery_percent": settings.mc_smf_battery_percent,
            "hr_battery_percent": settings.hr_battery_percent
        }},
        upsert=True
    )
    return {"message": "Recovery settings updated successfully"}

# Auth
@api_router.get("/users/list")
async def list_all_users():
    users = await db.users.find({}, {"_id": 0, "hashed_password": 0, "created_at": 0}).to_list(100)
    return users

@api_router.post("/auth/register", response_model=UserResponse)
async def register(user_data: UserCreate):
    existing_user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        name=user_data.name,
        email=user_data.email,
        hashed_password=hash_password(user_data.password)
    )
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    return UserResponse(id=user.id, name=user.name, email=user.email)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    access_token = create_access_token(data={"sub": user["id"]})
    return TokenResponse(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(id=user["id"], name=user["name"], email=user["email"])
    )

# Refining
@api_router.post("/refining/entries")
async def create_refining_entry(
    batches_data: str = Form(...),
    files: List[UploadFile] = File(...),
    entry_date: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    import json
    batches_json = json.loads(batches_data)
    
    file_idx = 0
    batches = []
    
    for batch_data in batches_json:
        batch = RefiningBatch(
            input_source=batch_data.get('input_source', 'manual'),
            sb_percentage=batch_data.get('sb_percentage'),
            lead_ingot_kg=batch_data['lead_ingot_kg'],
            lead_ingot_pieces=batch_data['lead_ingot_pieces'],
            lead_ingot_image=base64.b64encode(await files[file_idx].read()).decode('utf-8'),
            initial_dross_kg=batch_data['initial_dross_kg'],
            initial_dross_image=base64.b64encode(await files[file_idx + 1].read()).decode('utf-8'),
            dross_2nd_kg=batch_data['dross_2nd_kg'],
            dross_2nd_image=base64.b64encode(await files[file_idx + 2].read()).decode('utf-8'),
            dross_3rd_kg=batch_data['dross_3rd_kg'],
            dross_3rd_image=base64.b64encode(await files[file_idx + 3].read()).decode('utf-8'),
            dross_remarks=batch_data.get('dross_remarks', ''),
            pure_lead_kg=batch_data['pure_lead_kg'],
            pure_lead_pieces=batch_data.get('pure_lead_pieces', 0),
            pure_lead_image=base64.b64encode(await files[file_idx + 4].read()).decode('utf-8')
        )
        batches.append(batch)
        file_idx += 5
    
    entry = RefiningEntry(
        user_id=current_user["id"],
        user_name=current_user["name"],
        batches=batches
    )
    
    doc = entry.model_dump()
    
    # Handle custom entry date
    if entry_date:
        from datetime import datetime as dt
        custom_date = dt.fromisoformat(entry_date + "T12:00:00")
        doc['timestamp'] = custom_date.isoformat()
    else:
        doc['timestamp'] = doc['timestamp'].isoformat()
    
    for batch in doc['batches']:
        batch['timestamp'] = batch['timestamp'].isoformat()
    
    await db.entries.insert_one(doc)
    return {"id": entry.id, "message": "Refining entry created successfully"}

@api_router.delete("/admin/entries/{entry_id}")
async def delete_entry(entry_id: str, admin: dict = Depends(require_admin)):
    result = await db.entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted successfully"}

# Recycling
@api_router.post("/recycling/entries")
async def create_recycling_entry(
    batches_data: str = Form(...),
    files: List[UploadFile] = File(...),
    entry_date: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    import json
    batches_json = json.loads(batches_data)
    
    # Get recovery settings
    settings = await db.settings.find_one({"type": "recovery_settings"}, {"_id": 0})
    pp_percent = settings.get('pp_battery_percent', 60.5) / 100 if settings else 0.605
    mc_smf_percent = settings.get('mc_smf_battery_percent', 58.0) / 100 if settings else 0.58
    
    file_idx = 0
    batches = []
    
    for batch_data in batches_json:
        battery_kg = batch_data['battery_kg']
        battery_type = batch_data['battery_type']
        quantity_received = batch_data.get('quantity_received', 0)
        has_output_image = batch_data.get('has_output_image', False)
        
        if battery_type == "PP":
            remelted_lead_kg = battery_kg * pp_percent
        else:
            remelted_lead_kg = battery_kg * mc_smf_percent
        
        receivable_kg = remelted_lead_kg - quantity_received
        recovery_percent = (quantity_received / battery_kg * 100) if battery_kg > 0 else 0
        
        battery_image = base64.b64encode(await files[file_idx].read()).decode('utf-8')
        file_idx += 1
        
        if has_output_image:
            remelted_lead_image = base64.b64encode(await files[file_idx].read()).decode('utf-8')
            file_idx += 1
        else:
            remelted_lead_image = ""
        
        batch = RecyclingBatch(
            battery_type=battery_type,
            battery_kg=battery_kg,
            battery_image=battery_image,
            quantity_received=quantity_received,
            remelted_lead_kg=round(remelted_lead_kg, 2),
            receivable_kg=round(receivable_kg, 2),
            recovery_percent=round(recovery_percent, 2),
            remelted_lead_image=remelted_lead_image
        )
        batches.append(batch)
    
    entry = RecyclingEntry(
        user_id=current_user["id"],
        user_name=current_user["name"],
        batches=batches
    )
    
    doc = entry.model_dump()
    
    # Handle custom entry date
    if entry_date:
        from datetime import datetime as dt
        custom_date = dt.fromisoformat(entry_date + "T12:00:00")
        doc['timestamp'] = custom_date.isoformat()
    else:
        doc['timestamp'] = doc['timestamp'].isoformat()
    
    for batch in doc['batches']:
        batch['timestamp'] = batch['timestamp'].isoformat()
    
    await db.entries.insert_one(doc)
    return {"id": entry.id, "message": "Recycling entry created successfully"}

# Dross Recycling
@api_router.post("/dross-recycling/entries")
async def create_dross_recycling_entry(
    batches_data: str = Form(...),
    files: List[UploadFile] = File(...),
    entry_date: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    import json
    batches_json = json.loads(batches_data)
    
    batches = []
    for idx, batch_data in enumerate(batches_json):
        spectro_image = base64.b64encode(await files[idx].read()).decode('utf-8')
        
        batch = DrossRecyclingBatch(
            dross_type=batch_data['dross_type'],
            quantity_sent=batch_data['quantity_sent'],
            high_lead_recovered=batch_data['high_lead_recovered'],
            spectro_image=spectro_image
        )
        batches.append(batch)
    
    entry = DrossRecyclingEntry(
        user_id=current_user["id"],
        user_name=current_user["name"],
        batches=batches
    )
    
    doc = entry.model_dump()
    
    # Handle custom entry date
    if entry_date:
        from datetime import datetime as dt
        custom_date = dt.fromisoformat(entry_date + "T12:00:00")
        doc['timestamp'] = custom_date.isoformat()
    else:
        doc['timestamp'] = doc['timestamp'].isoformat()
    
    for batch in doc['batches']:
        batch['timestamp'] = batch['timestamp'].isoformat()
    
    await db.dross_recycling_entries.insert_one(doc)
    return {"id": entry.id, "message": "Dross recycling entry created successfully"}

@api_router.get("/dross-recycling/entries")
async def get_dross_recycling_entries(current_user: dict = Depends(get_current_user)):
    entries = await db.dross_recycling_entries.find({}, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    
    for entry in entries:
        if isinstance(entry['timestamp'], str):
            entry['timestamp'] = datetime.fromisoformat(entry['timestamp'])
        
        for batch in entry.get('batches', []):
            if isinstance(batch.get('timestamp'), str):
                batch['timestamp'] = datetime.fromisoformat(batch['timestamp'])
            batch.pop('spectro_image', None)
    
    return entries

@api_router.get("/dross-recycling/entries/{entry_id}")
async def get_dross_recycling_entry_detail(entry_id: str, current_user: dict = Depends(get_current_user)):
    entry = await db.dross_recycling_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    if isinstance(entry['timestamp'], str):
        entry['timestamp'] = datetime.fromisoformat(entry['timestamp'])
    
    for batch in entry.get('batches', []):
        if isinstance(batch.get('timestamp'), str):
            batch['timestamp'] = datetime.fromisoformat(batch['timestamp'])
    
    return entry

@api_router.delete("/admin/dross-recycling/{entry_id}")
async def delete_dross_recycling_entry(entry_id: str, admin: dict = Depends(require_admin)):
    result = await db.dross_recycling_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Dross recycling entry deleted successfully"}

# RML Purchases
@api_router.post("/rml-purchases")
async def create_rml_purchase(
    batches_data: str = Form(...),
    files: List[UploadFile] = File(...),
    entry_date: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    import json
    batches_json = json.loads(batches_data)
    
    batches = []
    for idx, batch_data in enumerate(batches_json):
        image = base64.b64encode(await files[idx].read()).decode('utf-8')
        
        # Generate SKU format: "remarks, sb%, date of inward"
        remarks = batch_data.get('remarks', 'RML')
        if not remarks:
            remarks = 'RML'
        
        # Use entry_date for the date of inward
        if entry_date:
            inward_date = entry_date  # Format: YYYY-MM-DD
            # Convert to DD/MM/YYYY format
            from datetime import datetime as dt
            date_obj = dt.fromisoformat(entry_date)
            formatted_date = date_obj.strftime("%d/%m/%Y")
        else:
            from datetime import datetime as dt
            formatted_date = dt.now().strftime("%d/%m/%Y")
        
        sku = f"{remarks}, {batch_data['sb_percentage']}%, {formatted_date}"
        
        batch = RMLPurchaseBatch(
            quantity_kg=batch_data['quantity_kg'],
            pieces=batch_data['pieces'],
            sb_percentage=batch_data['sb_percentage'],
            remarks=batch_data.get('remarks', ''),
            image=image,
            sku=sku
        )
        batches.append(batch)
    
    entry = RMLPurchaseEntry(
        user_id=current_user["id"],
        user_name=current_user["name"],
        batches=batches
    )
    
    doc = entry.model_dump()
    
    # Handle custom entry date
    if entry_date:
        from datetime import datetime as dt
        custom_date = dt.fromisoformat(entry_date + "T12:00:00")
        doc['timestamp'] = custom_date.isoformat()
    else:
        doc['timestamp'] = doc['timestamp'].isoformat()
    
    for batch in doc['batches']:
        batch['timestamp'] = batch['timestamp'].isoformat()
    
    await db.rml_purchases.insert_one(doc)
    return {"id": entry.id, "message": "RML purchase created successfully"}

@api_router.get("/rml-purchases")
async def get_rml_purchases(current_user: dict = Depends(get_current_user)):
    entries = await db.rml_purchases.find({}, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    
    for entry in entries:
        if isinstance(entry['timestamp'], str):
            entry['timestamp'] = datetime.fromisoformat(entry['timestamp'])
        
        for batch in entry.get('batches', []):
            if isinstance(batch.get('timestamp'), str):
                batch['timestamp'] = datetime.fromisoformat(batch['timestamp'])
            batch.pop('image', None)  # Don't return image data in list
    
    return entries

@api_router.get("/rml-purchases/skus")
async def get_rml_skus(current_user: dict = Depends(get_current_user)):
    """Get available RML SKUs for use in refining"""
    entries = await db.rml_purchases.find({}, {"_id": 0}).to_list(1000)
    
    # Aggregate SKUs with available quantities
    skus = {}
    for entry in entries:
        for batch in entry.get('batches', []):
            sku = batch.get('sku', '')
            if sku:
                if sku not in skus:
                    skus[sku] = {
                        'sku': sku,
                        'sb_percentage': batch.get('sb_percentage', 0),
                        'total_quantity_kg': 0,
                        'total_pieces': 0
                    }
                skus[sku]['total_quantity_kg'] += batch.get('quantity_kg', 0)
                skus[sku]['total_pieces'] += batch.get('pieces', 0)
    
    return list(skus.values())

@api_router.delete("/admin/rml-purchases/{entry_id}")
async def delete_rml_purchase(entry_id: str, admin: dict = Depends(require_admin)):
    result = await db.rml_purchases.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "RML purchase deleted successfully"}

@api_router.get("/entries")
async def get_entries(current_user: dict = Depends(get_current_user)):
    entries = await db.entries.find({}, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    
    for entry in entries:
        if isinstance(entry['timestamp'], str):
            entry['timestamp'] = datetime.fromisoformat(entry['timestamp'])
        
        for batch in entry.get('batches', []):
            if isinstance(batch.get('timestamp'), str):
                batch['timestamp'] = datetime.fromisoformat(batch['timestamp'])
            batch.pop('lead_ingot_image', None)
            batch.pop('initial_dross_image', None)
            batch.pop('dross_2nd_image', None)
            batch.pop('dross_3rd_image', None)
            batch.pop('pure_lead_image', None)
            batch.pop('battery_image', None)
            batch.pop('remelted_lead_image', None)
    
    return entries

@api_router.get("/entries/{entry_id}")
async def get_entry_detail(entry_id: str, current_user: dict = Depends(get_current_user)):
    entry = await db.entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    if isinstance(entry['timestamp'], str):
        entry['timestamp'] = datetime.fromisoformat(entry['timestamp'])
    
    for batch in entry.get('batches', []):
        if isinstance(batch.get('timestamp'), str):
            batch['timestamp'] = datetime.fromisoformat(batch['timestamp'])
    
    return entry

@api_router.get("/dross")
async def get_dross_data(current_user: dict = Depends(get_current_user)):
    refining_entries = await db.entries.find({"entry_type": "refining"}, {"_id": 0}).to_list(10000)
    
    dross_data = []
    for entry in refining_entries:
        if isinstance(entry['timestamp'], str):
            entry['timestamp'] = datetime.fromisoformat(entry['timestamp'])
        
        for batch_idx, batch in enumerate(entry.get('batches', []), 1):
            if isinstance(batch.get('timestamp'), str):
                batch['timestamp'] = datetime.fromisoformat(batch['timestamp'])
            
            dross_data.append({
                'entry_id': entry['id'],
                'user_name': entry['user_name'],
                'timestamp': entry['timestamp'],
                'batch_number': batch_idx,
                'initial_dross_kg': batch['initial_dross_kg'],
                'dross_2nd_kg': batch['dross_2nd_kg'],
                'dross_3rd_kg': batch['dross_3rd_kg'],
                'total_dross': batch['initial_dross_kg'] + batch['dross_2nd_kg'] + batch['dross_3rd_kg']
            })
    
    dross_data.sort(key=lambda x: x['timestamp'], reverse=True)
    return dross_data

# Sales
@api_router.post("/sales", response_model=SaleEntry)
async def create_sale(sale_data: SaleCreate, current_user: dict = Depends(get_current_user)):
    sale = SaleEntry(
        user_id=current_user["id"],
        user_name=current_user["name"],
        party_name=sale_data.party_name,
        sku_type=sale_data.sku_type,
        quantity_kg=sale_data.quantity_kg,
        entry_date=sale_data.entry_date
    )
    
    doc = sale.model_dump()
    
    # Handle custom entry date
    if sale_data.entry_date:
        from datetime import datetime as dt
        custom_date = dt.fromisoformat(sale_data.entry_date + "T12:00:00")
        doc['timestamp'] = custom_date.isoformat()
    else:
        doc['timestamp'] = doc['timestamp'].isoformat()
    
    await db.sales.insert_one(doc)
    
    return sale

@api_router.get("/sales/available-skus", response_model=List[AvailableSKU])
async def get_available_skus(current_user: dict = Depends(get_current_user)):
    """Get all available SKUs with their current stock for sales"""
    available_skus = []
    
    # Get all refining entries for calculations
    refining_entries = await db.entries.find({"entry_type": "refining"}, {"_id": 0}).to_list(10000)
    
    # Get all sales
    sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    
    # Calculate Pure Lead stock
    total_pure_lead = sum(
        batch['pure_lead_kg']
        for entry in refining_entries
        for batch in entry.get('batches', [])
    )
    pure_lead_sold = sum(sale['quantity_kg'] for sale in sales if sale.get('sku_type') == 'Pure Lead')
    pure_lead_stock = max(0, total_pure_lead - pure_lead_sold)
    
    if pure_lead_stock > 0:
        available_skus.append(AvailableSKU(
            sku_type="Pure Lead",
            sb_percentage=None,
            available_kg=round(pure_lead_stock, 2),
            display_name="Pure Lead"
        ))
    
    # Calculate High Lead stock
    dross_recycling_entries = await db.dross_recycling_entries.find({}, {"_id": 0}).to_list(10000)
    total_high_lead = sum(
        batch['high_lead_recovered']
        for entry in dross_recycling_entries
        for batch in entry.get('batches', [])
    )
    high_lead_sold = sum(sale['quantity_kg'] for sale in sales if sale.get('sku_type') == 'High Lead')
    high_lead_stock = max(0, total_high_lead - high_lead_sold)
    
    if high_lead_stock > 0:
        available_skus.append(AvailableSKU(
            sku_type="High Lead",
            sb_percentage=None,
            available_kg=round(high_lead_stock, 2),
            display_name="High Lead"
        ))
    
    # Calculate RML SKUs stock
    rml_purchases = await db.rml_purchases.find({}, {"_id": 0}).to_list(10000)
    
    # Aggregate RML purchases by SKU
    rml_skus = {}
    for entry in rml_purchases:
        for batch in entry.get('batches', []):
            sku = batch.get('sku', '')
            if sku:
                if sku not in rml_skus:
                    rml_skus[sku] = {
                        'purchased_kg': 0,
                        'sb_percentage': batch.get('sb_percentage', 0)
                    }
                rml_skus[sku]['purchased_kg'] += batch.get('quantity_kg', 0)
    
    # Deduct RML used in refining
    for entry in refining_entries:
        for batch in entry.get('batches', []):
            input_source = batch.get('input_source', 'manual')
            if input_source != 'manual' and input_source != 'SANTOSH':
                # This is an RML SKU used in refining
                if input_source in rml_skus:
                    rml_skus[input_source]['purchased_kg'] -= batch.get('lead_ingot_kg', 0)
    
    # Deduct RML sold
    for sale in sales:
        sku_type = sale.get('sku_type', '')
        if sku_type and sku_type not in ['Pure Lead', 'High Lead']:
            if sku_type in rml_skus:
                rml_skus[sku_type]['purchased_kg'] -= sale['quantity_kg']
    
    # Add RML SKUs with positive stock
    for sku, data in rml_skus.items():
        stock = max(0, data['purchased_kg'])
        if stock > 0:
            available_skus.append(AvailableSKU(
                sku_type=sku,
                sb_percentage=data['sb_percentage'],
                available_kg=round(stock, 2),
                display_name=f"{sku} (SB: {data['sb_percentage']}%)"
            ))
    
    return available_skus

@api_router.get("/sales")
async def get_sales(current_user: dict = Depends(get_current_user)):
    sales = await db.sales.find({}, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    
    for sale in sales:
        if isinstance(sale['timestamp'], str):
            sale['timestamp'] = datetime.fromisoformat(sale['timestamp'])
    
    return sales

@api_router.delete("/admin/sales/{sale_id}")
async def delete_sale(sale_id: str, admin: dict = Depends(require_admin)):
    result = await db.sales.delete_one({"id": sale_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sale not found")
    return {"message": "Sale deleted successfully"}

# Summary
@api_router.get("/summary", response_model=SummaryStats)
async def get_summary(current_user: dict = Depends(get_current_user)):
    # Get all refining entries
    refining_entries = await db.entries.find({"entry_type": "refining"}, {"_id": 0}).to_list(10000)
    
    # Calculate total pure lead produced from refining
    total_pure_lead_manufactured = sum(
        batch['pure_lead_kg']
        for entry in refining_entries
        for batch in entry.get('batches', [])
    )
    
    # Calculate total dross from refining
    total_dross = sum(
        batch['initial_dross_kg'] + batch['dross_2nd_kg'] + batch['dross_3rd_kg']
        for entry in refining_entries
        for batch in entry.get('batches', [])
    )
    
    # Calculate Antimony Recoverable from Drosses
    # Formula: SB% x Quantity / 100 for each batch (to get actual antimony weight)
    antimony_recoverable = sum(
        (batch.get('sb_percentage', 0) or 0) * batch.get('lead_ingot_kg', 0) / 100
        for entry in refining_entries
        for batch in entry.get('batches', [])
        if batch.get('sb_percentage')  # Only count batches with SB% entered
    )
    
    # Get dross recycling entries for High Lead
    dross_recycling_entries = await db.dross_recycling_entries.find({}, {"_id": 0}).to_list(10000)
    total_high_lead = sum(
        batch['high_lead_recovered']
        for entry in dross_recycling_entries
        for batch in entry.get('batches', [])
    )
    
    # Get recycling entries for receivable
    recycling_entries = await db.entries.find({"entry_type": "recycling"}, {"_id": 0}).to_list(10000)
    total_remelted_lead = sum(
        batch.get('quantity_received', 0)
        for entry in recycling_entries
        for batch in entry.get('batches', [])
    )
    
    # Calculate lead receivable from recycling
    raw_receivable = sum(
        batch.get('receivable_kg', 0)
        for entry in recycling_entries
        for batch in entry.get('batches', [])
    )
    
    # Calculate SANTOSH usage (deduct from receivable)
    santosh_usage = sum(
        batch.get('lead_ingot_kg', 0)
        for entry in refining_entries
        for batch in entry.get('batches', [])
        if batch.get('input_source') == 'SANTOSH'
    )
    
    # Adjust receivable by deducting SANTOSH usage
    total_receivable = max(0, raw_receivable - santosh_usage)
    
    # RML Purchases
    rml_purchases = await db.rml_purchases.find({}, {"_id": 0}).to_list(10000)
    total_rml_purchased = sum(
        batch.get('quantity_kg', 0)
        for entry in rml_purchases
        for batch in entry.get('batches', [])
    )
    
    # Calculate RML used in refining (non-manual, non-SANTOSH input sources)
    rml_used_in_refining = sum(
        batch.get('lead_ingot_kg', 0)
        for entry in refining_entries
        for batch in entry.get('batches', [])
        if batch.get('input_source') not in ['manual', 'SANTOSH', None, '']
    )
    
    # Get all sales
    sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    total_sold = sum(sale['quantity_kg'] for sale in sales)
    
    # Calculate SKU-specific sales
    pure_lead_sold = sum(sale['quantity_kg'] for sale in sales if sale.get('sku_type') == 'Pure Lead')
    high_lead_sold = sum(sale['quantity_kg'] for sale in sales if sale.get('sku_type') == 'High Lead')
    rml_sold = sum(sale['quantity_kg'] for sale in sales if sale.get('sku_type') not in ['Pure Lead', 'High Lead', None, ''])
    
    # Calculate final stocks
    pure_lead_stock = max(0, total_pure_lead_manufactured - pure_lead_sold)
    rml_stock = max(0, total_rml_purchased - rml_used_in_refining - rml_sold)
    high_lead_stock = max(0, total_high_lead - high_lead_sold)
    
    # Legacy calculations for backward compatibility
    remelted_lead_in_stock = max(0, total_remelted_lead + total_rml_purchased - total_sold)
    available_stock = pure_lead_stock + rml_stock + high_lead_stock
    
    return SummaryStats(
        # New simplified dashboard stats
        pure_lead_stock=round(pure_lead_stock, 2),
        rml_stock=round(rml_stock, 2),
        total_receivable=round(total_receivable, 2),
        high_lead_stock=round(high_lead_stock, 2),
        total_dross=round(total_dross, 2),
        antimony_recoverable=round(antimony_recoverable, 2),
        # Legacy fields
        total_pure_lead_manufactured=round(total_pure_lead_manufactured, 2),
        total_remelted_lead=round(total_remelted_lead, 2),
        total_sold=round(total_sold, 2),
        available_stock=round(available_stock, 2),
        remelted_lead_in_stock=round(remelted_lead_in_stock, 2),
        total_high_lead=round(total_high_lead, 2),
        total_rml_purchased=round(total_rml_purchased, 2)
    )

@api_router.get("/dross/export/excel")
async def export_dross_excel(current_user: dict = Depends(get_current_user)):
    refining_entries = await db.entries.find({"entry_type": "refining"}, {"_id": 0}).sort("timestamp", -1).to_list(10000)
    
    wb = Workbook()
    ws_dross = wb.active
    ws_dross.title = "Dross Data"
    
    headers_dross = [
        "Date", "Time", "Employee", "Batch #",
        "Initial Dross (kg)", "2nd Dross (kg)", "3rd Dross (kg)", "Total Dross (kg)"
    ]
    
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers_dross, 1):
        cell = ws_dross.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row_num = 2
    for entry in refining_entries:
        timestamp = entry['timestamp']
        if isinstance(timestamp, str):
            timestamp = datetime.fromisoformat(timestamp)
        
        for batch_idx, batch in enumerate(entry.get('batches', []), 1):
            total_dross = batch['initial_dross_kg'] + batch['dross_2nd_kg'] + batch['dross_3rd_kg']
            
            ws_dross.cell(row=row_num, column=1, value=timestamp.strftime("%Y-%m-%d"))
            ws_dross.cell(row=row_num, column=2, value=timestamp.strftime("%H:%M:%S"))
            ws_dross.cell(row=row_num, column=3, value=entry['user_name'])
            ws_dross.cell(row=row_num, column=4, value=f"Batch {batch_idx}")
            ws_dross.cell(row=row_num, column=5, value=batch['initial_dross_kg'])
            ws_dross.cell(row=row_num, column=6, value=batch['dross_2nd_kg'])
            ws_dross.cell(row=row_num, column=7, value=batch['dross_3rd_kg'])
            ws_dross.cell(row=row_num, column=8, value=total_dross)
            row_num += 1
    
    dross_recycling = await db.dross_recycling_entries.find({}, {"_id": 0}).sort("timestamp", -1).to_list(10000)
    if dross_recycling:
        ws_high_lead = wb.create_sheet("HIGH LEAD Recovery")
        headers_high = [
            "Date", "Time", "Employee",
            "Dross Type", "Quantity Sent (kg)", "HIGH LEAD Recovered (kg)"
        ]
        
        for col_num, header in enumerate(headers_high, 1):
            cell = ws_high_lead.cell(row=1, column=col_num, value=header)
            cell.fill = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num = 2
        for entry in dross_recycling:
            timestamp = entry['timestamp']
            if isinstance(timestamp, str):
                timestamp = datetime.fromisoformat(timestamp)
            
            for batch in entry.get('batches', []):
                ws_high_lead.cell(row=row_num, column=1, value=timestamp.strftime("%Y-%m-%d"))
                ws_high_lead.cell(row=row_num, column=2, value=timestamp.strftime("%H:%M:%S"))
                ws_high_lead.cell(row=row_num, column=3, value=entry['user_name'])
                ws_high_lead.cell(row=row_num, column=4, value=batch['dross_type'].upper())
                ws_high_lead.cell(row=row_num, column=5, value=batch['quantity_sent'])
                ws_high_lead.cell(row=row_num, column=6, value=batch['high_lead_recovered'])
                row_num += 1
    
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dross_data.xlsx"}
    )

@api_router.get("/entries/export/excel")
async def export_entries_excel(current_user: dict = Depends(get_current_user)):
    entries = await db.entries.find({}, {"_id": 0}).sort("timestamp", -1).to_list(10000)
    
    wb = Workbook()
    
    ws_refining = wb.active
    ws_refining.title = "Refining"
    
    headers_refining = [
        "Date", "Time", "Employee", "Batch #",
        "Lead Ingot (kg)", "Pieces",
        "Initial Dross (kg)", "2nd Dross (kg)", "3rd Dross (kg)",
        "Pure Lead Output (kg)"
    ]
    
    header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers_refining, 1):
        cell = ws_refining.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row_num = 2
    for entry in entries:
        if entry.get('entry_type') == 'refining':
            timestamp = entry['timestamp']
            if isinstance(timestamp, str):
                timestamp = datetime.fromisoformat(timestamp)
            
            for batch_idx, batch in enumerate(entry.get('batches', []), 1):
                ws_refining.cell(row=row_num, column=1, value=timestamp.strftime("%Y-%m-%d"))
                ws_refining.cell(row=row_num, column=2, value=timestamp.strftime("%H:%M:%S"))
                ws_refining.cell(row=row_num, column=3, value=entry['user_name'])
                ws_refining.cell(row=row_num, column=4, value=f"Batch {batch_idx}")
                ws_refining.cell(row=row_num, column=5, value=batch['lead_ingot_kg'])
                ws_refining.cell(row=row_num, column=6, value=batch['lead_ingot_pieces'])
                ws_refining.cell(row=row_num, column=7, value=batch['initial_dross_kg'])
                ws_refining.cell(row=row_num, column=8, value=batch['dross_2nd_kg'])
                ws_refining.cell(row=row_num, column=9, value=batch['dross_3rd_kg'])
                ws_refining.cell(row=row_num, column=10, value=batch['pure_lead_kg'])
                row_num += 1
    
    ws_recycling = wb.create_sheet("Recycling")
    headers_recycling = [
        "Date", "Time", "Employee", "Batch #",
        "Battery Type", "Battery Input (kg)", "Expected Output (kg)", 
        "Quantity Received (kg)", "Receivable (kg)", "Recovery %"
    ]
    
    for col_num, header in enumerate(headers_recycling, 1):
        cell = ws_recycling.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row_num = 2
    for entry in entries:
        if entry.get('entry_type') == 'recycling':
            timestamp = entry['timestamp']
            if isinstance(timestamp, str):
                timestamp = datetime.fromisoformat(timestamp)
            
            for batch_idx, batch in enumerate(entry.get('batches', []), 1):
                ws_recycling.cell(row=row_num, column=1, value=timestamp.strftime("%Y-%m-%d"))
                ws_recycling.cell(row=row_num, column=2, value=timestamp.strftime("%H:%M:%S"))
                ws_recycling.cell(row=row_num, column=3, value=entry['user_name'])
                ws_recycling.cell(row=row_num, column=4, value=f"Batch {batch_idx}")
                ws_recycling.cell(row=row_num, column=5, value=batch['battery_type'])
                ws_recycling.cell(row=row_num, column=6, value=batch['battery_kg'])
                ws_recycling.cell(row=row_num, column=7, value=batch['remelted_lead_kg'])
                ws_recycling.cell(row=row_num, column=8, value=batch.get('quantity_received', 0))
                ws_recycling.cell(row=row_num, column=9, value=batch.get('receivable_kg', 0))
                ws_recycling.cell(row=row_num, column=10, value=batch.get('recovery_percent', 0))
                row_num += 1
    
    ws_sales = wb.create_sheet("Sales")
    headers_sales = ["Date", "Time", "Employee", "Party Name", "Quantity Sold (kg)"]
    
    for col_num, header in enumerate(headers_sales, 1):
        cell = ws_sales.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    sales = await db.sales.find({}, {"_id": 0}).sort("timestamp", -1).to_list(10000)
    for row_num, sale in enumerate(sales, 2):
        timestamp = sale['timestamp']
        if isinstance(timestamp, str):
            timestamp = datetime.fromisoformat(timestamp)
        
        ws_sales.cell(row=row_num, column=1, value=timestamp.strftime("%Y-%m-%d"))
        ws_sales.cell(row=row_num, column=2, value=timestamp.strftime("%H:%M:%S"))
        ws_sales.cell(row=row_num, column=3, value=sale['user_name'])
        ws_sales.cell(row=row_num, column=4, value=sale['party_name'])
        ws_sales.cell(row=row_num, column=5, value=sale['quantity_kg'])
    
    for ws in [ws_refining, ws_recycling, ws_sales]:
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leadtrack_report.xlsx"}
    )

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()